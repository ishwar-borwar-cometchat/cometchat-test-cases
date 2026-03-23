import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def style_rows(ws, start_row, end_row):
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap = Alignment(wrap_text=True, vertical='top')
    for row in range(start_row, end_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = wrap
            cell.border = thin_border

# ============================================================
# PATCH SEND MESSAGE - Add Custom Message & Audio test cases
# ============================================================
wb = openpyxl.load_workbook('cometchat-test-cases/Send_Message/Send_Message_Test_Cases.xlsx')
ws = wb.active
start_row = ws.max_row + 1

send_msg_additions = [
    # --- Custom Message ---
    ["MSG_094", "Positive", "Verify Custom Message", "User is logged in and chat is open",
     "Verify sending a custom message (poll)",
     "1. Open a conversation.\n2. Create a poll via custom message option.\n3. Send.",
     "Poll should be sent and display in chat with options for recipients to vote.",
     "High / High"],
    ["MSG_095", "", "", "",
     "Verify sending a custom message (meeting link)",
     "1. Open a conversation.\n2. Send a meeting/conference link as custom message.",
     "Custom message with meeting link should display with appropriate formatting and join button.",
     "Medium / Medium"],
    ["MSG_096", "", "", "",
     "Verify sending a custom message (collaborative document)",
     "1. Open a conversation.\n2. Share a collaborative document as custom message.",
     "Custom message should display with document preview and open/join action.",
     "Medium / Medium"],
    ["MSG_097", "", "", "",
     "Verify custom message renders correctly for receiver",
     "1. Send a custom message from User A.\n2. Observe on User B's side.",
     "Custom message should render correctly with all interactive elements on receiver's end.",
     "High / High"],
    ["MSG_098", "", "", "",
     "Verify custom message in group chat",
     "1. Open a group conversation.\n2. Send a custom message.",
     "Custom message should display correctly for all group members.",
     "High / High"],
    ["MSG_099", "Negative", "Verify Custom Message", "User is logged in, network disconnected",
     "Verify custom message fails without network",
     "1. Disconnect network.\n2. Try to send a custom message.",
     "Error message should display or message should queue with retry option.",
     "High / High"],

    # --- Audio Message ---
    ["MSG_100", "Positive", "Verify Audio Message", "User is logged in and chat is open",
     "Verify sending an audio file attachment",
     "1. Open a conversation.\n2. Click attach.\n3. Select an audio file (.mp3, .wav).\n4. Send.",
     "Audio file should upload and display with audio player, file name, and duration.",
     "High / High"],
    ["MSG_101", "", "", "",
     "Verify audio message playback by sender",
     "1. Send an audio message.\n2. Click play on the audio player.",
     "Audio should play with progress bar, play/pause button, and duration indicator.",
     "High / High"],
    ["MSG_102", "", "", "",
     "Verify audio message playback by receiver",
     "1. Receive an audio message.\n2. Click play.",
     "Audio should download (if needed) and play correctly with player controls.",
     "High / High"],
    ["MSG_103", "", "", "",
     "Verify audio message shows duration",
     "1. Send/receive an audio message.\n2. Observe audio player.",
     "Audio duration should display (e.g., '0:45', '2:30').",
     "Medium / Medium"],
    ["MSG_104", "", "", "",
     "Verify sending recorded audio message",
     "1. Open a conversation.\n2. Press and hold record button.\n3. Record audio.\n4. Release to send.",
     "Recorded audio should be sent and display with audio player in chat.",
     "High / High"],
    ["MSG_105", "", "", "",
     "Verify audio message in group chat",
     "1. Open a group conversation.\n2. Send an audio message.",
     "Audio message should display correctly for all group members with sender info.",
     "High / High"],
    ["MSG_106", "Negative", "Verify Audio Message", "User is logged in and chat is open",
     "Verify large audio file exceeding size limit",
     "1. Try to send an audio file exceeding the size limit.",
     "Error message should display indicating file size limit exceeded.",
     "High / Medium"],
    ["MSG_107", "", "", "User is logged in, network disconnected",
     "Verify audio message fails without network",
     "1. Disconnect network.\n2. Try to send an audio message.",
     "Error message should display or upload should queue with retry option.",
     "High / High"],
    ["MSG_108", "", "", "User is logged in and chat is open",
     "Verify unsupported audio format",
     "1. Try to send an unsupported audio file format.",
     "Error message should display or file should be rejected.",
     "Medium / Medium"],
]

for row_data in send_msg_additions:
    ws.append(row_data)

style_rows(ws, start_row, ws.max_row)
wb.save('cometchat-test-cases/Send_Message/Send_Message_Test_Cases.xlsx')
print(f"Send Message: Added {len(send_msg_additions)} test cases (MSG_094 to MSG_108). Total rows: {ws.max_row}")

# ============================================================
# PATCH CALL MODULE - Add Audio Call & Incoming call test cases
# ============================================================
wb2 = openpyxl.load_workbook('cometchat-test-cases/Call_Module/Calls_Module_Test_Cases.xlsx')
ws2 = wb2.active
start_row2 = ws2.max_row + 1

call_additions = [
    # --- Audio Call ---
    ["CALLS_043", "Positive", "Verify Audio Call", "User is logged in and on a conversation",
     "Verify initiating an audio call",
     "1. Open a one-on-one conversation.\n2. Click the audio call (phone) icon in header.",
     "Audio call should initiate. Calling screen should appear with user name, avatar, and call controls.",
     "High / High"],
    ["CALLS_044", "", "", "",
     "Verify audio call ringing state",
     "1. Initiate an audio call to another user.",
     "Ringing state should display with 'Calling...' or ringing indicator. Ring tone should play.",
     "High / High"],
    ["CALLS_045", "", "", "",
     "Verify audio call connected state",
     "1. Initiate audio call.\n2. Other user accepts.",
     "Call should connect. Timer should start. Audio should be active both ways.",
     "High / High"],
    ["CALLS_046", "", "", "",
     "Verify audio call mute/unmute",
     "1. During an active audio call.\n2. Click mute button.\n3. Click unmute.",
     "Mute should silence outgoing audio. Unmute should restore it. Icon should toggle.",
     "High / High"],
    ["CALLS_047", "", "", "",
     "Verify audio call speaker toggle",
     "1. During an active audio call.\n2. Click speaker button.",
     "Audio should switch between earpiece and speaker. Icon should toggle.",
     "Medium / Medium"],
    ["CALLS_048", "", "", "",
     "Verify ending an audio call",
     "1. During an active audio call.\n2. Click end call (red) button.",
     "Call should end. Both users should return to chat. Call duration should log.",
     "High / High"],
    ["CALLS_049", "", "", "",
     "Verify audio call in group",
     "1. Open a group conversation.\n2. Initiate an audio call.",
     "Group audio call should initiate. All members should receive call notification.",
     "High / High"],
    ["CALLS_050", "Negative", "Verify Audio Call", "User is logged in",
     "Verify audio call when user is offline",
     "1. Try to audio call a user who is offline.",
     "Call should ring and eventually show 'User unavailable' or similar message.",
     "High / High"],
    ["CALLS_051", "", "", "User is logged in, network disconnected",
     "Verify audio call fails without network",
     "1. Disconnect network.\n2. Try to initiate an audio call.",
     "Error message should display indicating no network connection.",
     "High / High"],
    ["CALLS_052", "", "", "User is logged in and on an active call",
     "Verify audio call drops on network loss",
     "1. Be on an active audio call.\n2. Disconnect network.",
     "Call should drop with reconnecting attempt or error message.",
     "High / High"],

    # --- Incoming Call ---
    ["CALLS_053", "Positive", "Verify Incoming Call", "User is logged in and app is open",
     "Verify incoming audio call notification",
     "1. Another user initiates an audio call to you.",
     "Incoming call screen should appear with caller name, avatar, accept and reject buttons.",
     "High / High"],
    ["CALLS_054", "", "", "",
     "Verify incoming video call notification",
     "1. Another user initiates a video call to you.",
     "Incoming call screen should appear with caller name, avatar, accept and reject buttons.",
     "High / High"],
    ["CALLS_055", "", "", "",
     "Verify accepting an incoming audio call",
     "1. Receive an incoming audio call.\n2. Click accept/answer button.",
     "Audio call should connect. Call screen should show with active call controls.",
     "High / High"],
    ["CALLS_056", "", "", "",
     "Verify accepting an incoming video call",
     "1. Receive an incoming video call.\n2. Click accept/answer button.",
     "Video call should connect with both video feeds visible.",
     "High / High"],
    ["CALLS_057", "", "", "",
     "Verify rejecting an incoming call",
     "1. Receive an incoming call.\n2. Click reject/decline button.",
     "Call should be rejected. Caller should see 'Call declined' or similar. Call logged as missed.",
     "High / High"],
    ["CALLS_058", "", "", "",
     "Verify incoming call when app is in background",
     "1. Minimize the app.\n2. Another user calls you.",
     "Push notification or system call UI should appear allowing accept/reject.",
     "High / High"],
    ["CALLS_059", "", "", "",
     "Verify incoming group call notification",
     "1. A group member initiates a group call.",
     "Incoming group call notification should appear with group name and accept/reject options.",
     "High / High"],
    ["CALLS_060", "", "", "",
     "Verify missed call when not answered",
     "1. Receive an incoming call.\n2. Do not answer until it times out.",
     "Call should be logged as 'Missed' in call logs. Missed call notification should appear.",
     "High / High"],
    ["CALLS_061", "Negative", "Verify Incoming Call", "User is logged in with Do Not Disturb enabled",
     "Verify incoming call behavior with DND",
     "1. Enable Do Not Disturb.\n2. Receive an incoming call.",
     "Call should be silenced or blocked based on DND settings. Call logged as missed.",
     "Medium / Medium"],
    ["CALLS_062", "", "", "User is logged in and already on a call",
     "Verify incoming call while on another call",
     "1. Be on an active call.\n2. Receive another incoming call.",
     "Second call notification should appear with option to accept (end current) or reject.",
     "High / High"],
]

for row_data in call_additions:
    ws2.append(row_data)

style_rows(ws2, start_row2, ws2.max_row)
wb2.save('cometchat-test-cases/Call_Module/Calls_Module_Test_Cases.xlsx')
print(f"Call Module: Added {len(call_additions)} test cases (CALLS_043 to CALLS_062). Total rows: {ws2.max_row}")

print("\nAll patches applied successfully!")
