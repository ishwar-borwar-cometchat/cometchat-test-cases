from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Send Message Test Cases"

# Headers
headers = ["TestCase_ID", "Sentiment", "Test Scenario", "Precondition", "Test Case", "Steps", "Expected Results", "Priority / Severity"]

# Data for Send Message module
data = [
    # Verify Message Input field - Positive
    ["MSG_001", "Positive", "Verify Message Input field", "User is logged in and chat is open", "Verify message input field is visible", "1. Open any chat conversation.\n2. Observe bottom of chat window.", "Message input field should be visible with placeholder text.", "High / High"],
    ["MSG_002", "Positive", "Verify Message Input field", "User is logged in and chat is open", "Verify message input field is clickable", "1. Open any chat conversation.\n2. Click on message input field.", "Input field should become active; cursor should appear; keyboard should open on mobile.", "High / High"],
    ["MSG_003", "Positive", "Verify Message Input field", "User is logged in and chat is open", "Verify typing in message input field", "1. Open any chat conversation.\n2. Click on input field.\n3. Type a message.", "Typed text should appear in the input field.", "High / High"],
    ["MSG_004", "Positive", "Verify Message Input field", "User is logged in and chat is open", "Verify multi-line message input", "1. Open any chat conversation.\n2. Type a long message.\n3. Observe input field expansion.", "Input field should expand or allow scrolling for multi-line messages.", "Medium / Medium"],
    # Verify Message Input field - Negative
    ["MSG_005", "Negative", "Verify Message Input field", "User is logged in and chat is open", "Verify empty message cannot be sent", "1. Open any chat conversation.\n2. Leave input field empty.\n3. Click send button.", "Send button should be disabled or message should not be sent.", "High / High"],
    ["MSG_006", "Negative", "Verify Message Input field", "User is logged in and chat is open", "Verify message with only spaces", "1. Open any chat conversation.\n2. Type only spaces.\n3. Click send button.", "Message should not be sent; spaces-only input should be treated as empty.", "Medium / Medium"],
    
    # Verify Send button - Positive
    ["MSG_007", "Positive", "Verify Send button", "User is logged in and chat is open", "Verify send button is visible", "1. Open any chat conversation.\n2. Observe send button next to input field.", "Send button (arrow/paper plane icon) should be visible.", "High / High"],
    ["MSG_008", "Positive", "Verify Send button", "User is logged in and chat is open", "Verify send button enabled when text entered", "1. Open any chat conversation.\n2. Type a message.\n3. Observe send button.", "Send button should become enabled/highlighted when text is entered.", "High / High"],
    ["MSG_009", "Positive", "Verify Send button", "User is logged in and message typed", "Verify send button click sends message", "1. Type a message.\n2. Click send button.\n3. Observe chat.", "Message should be sent and appear in chat with sent status.", "High / High"],
    ["MSG_010", "Positive", "Verify Send button", "User is logged in and message typed", "Verify send button visual feedback on click", "1. Type a message.\n2. Click send button.\n3. Observe button.", "Send button should show click feedback (color change, animation).", "Low / Low"],
    # Verify Send button - Negative
    ["MSG_011", "Negative", "Verify Send button", "User is logged in and chat is open", "Verify send button disabled when empty", "1. Open any chat conversation.\n2. Leave input field empty.\n3. Observe send button.", "Send button should be disabled or grayed out.", "High / Medium"],
    
    # Verify Text message sending - Positive
    ["MSG_012", "Positive", "Verify Text message sending", "User is logged in and chat is open", "Verify sending simple text message", "1. Open any chat conversation.\n2. Type \"Hello\".\n3. Click send.", "Message \"Hello\" should appear in chat on right side (sent).", "High / High"],
    ["MSG_013", "Positive", "Verify Text message sending", "User is logged in and chat is open", "Verify sending long text message", "1. Open any chat conversation.\n2. Type a 500+ character message.\n3. Click send.", "Long message should be sent successfully and display properly.", "Medium / Medium"],
    ["MSG_014", "Positive", "Verify Text message sending", "User is logged in and chat is open", "Verify sending message with special characters", "1. Open any chat conversation.\n2. Type \"Hello @#$%^&*()!\".\n3. Click send.", "Message with special characters should be sent and displayed correctly.", "Medium / Medium"],
    ["MSG_015", "Positive", "Verify Text message sending", "User is logged in and chat is open", "Verify sending message with emojis", "1. Open any chat conversation.\n2. Type \"Hello 😀🎉👍\".\n3. Click send.", "Message with emojis should be sent and emojis displayed correctly.", "Medium / Medium"],
    ["MSG_016", "Positive", "Verify Text message sending", "User is logged in and chat is open", "Verify sending message with numbers", "1. Open any chat conversation.\n2. Type \"Order #12345\".\n3. Click send.", "Message with numbers should be sent correctly.", "Low / Low"],
    ["MSG_017", "Positive", "Verify Text message sending", "User is logged in and chat is open", "Verify sending message with URL", "1. Open any chat conversation.\n2. Type \"Check https://example.com\".\n3. Click send.", "Message should be sent; URL may be clickable/highlighted.", "Medium / Medium"],
    # Verify Text message sending - Negative
    ["MSG_018", "Negative", "Verify Text message sending", "User is logged in, network disconnected", "Verify message sending fails without network", "1. Disconnect network.\n2. Type a message.\n3. Click send.", "Error indicator should show; message should be queued or show failed status.", "High / High"],
    ["MSG_019", "Negative", "Verify Text message sending", "User is logged in and chat is open", "Verify extremely long message handling", "1. Type a 10000+ character message.\n2. Click send.", "Application should handle gracefully; either send or show character limit warning.", "Low / Medium"],
    
    # Verify Sent message display - Positive
    ["MSG_020", "Positive", "Verify Sent message display", "User has sent a message", "Verify sent message alignment", "1. Send a message.\n2. Observe message position in chat.", "Sent message should appear on the right side of chat window.", "High / High"],
    ["MSG_021", "Positive", "Verify Sent message display", "User has sent a message", "Verify sent message bubble color", "1. Send a message.\n2. Observe message bubble.", "Sent message should have distinct bubble color (e.g., blue/purple).", "Medium / Medium"],
    ["MSG_022", "Positive", "Verify Sent message display", "User has sent a message", "Verify sent message timestamp", "1. Send a message.\n2. Observe timestamp.", "Timestamp should display below or next to sent message.", "Medium / Medium"],
    ["MSG_023", "Positive", "Verify Sent message display", "User has sent a message", "Verify sent message status indicator", "1. Send a message.\n2. Observe status indicator.", "Status indicator should show (single tick = sent, double tick = delivered, blue tick = read).", "High / High"],
    
    # Verify Received message display - Positive
    ["MSG_024", "Positive", "Verify Received message display", "User has received a message", "Verify received message alignment", "1. Receive a message from another user.\n2. Observe message position.", "Received message should appear on the left side of chat window.", "High / High"],
    ["MSG_025", "Positive", "Verify Received message display", "User has received a message", "Verify received message bubble color", "1. Receive a message.\n2. Observe message bubble.", "Received message should have different bubble color than sent (e.g., gray/white).", "Medium / Medium"],
    ["MSG_026", "Positive", "Verify Received message display", "User has received a message", "Verify received message sender info", "1. Receive a message in group chat.\n2. Observe sender info.", "Sender name and/or avatar should display with received message in group chats.", "Medium / Medium"],
    ["MSG_027", "Positive", "Verify Received message display", "User has received a message", "Verify received message timestamp", "1. Receive a message.\n2. Observe timestamp.", "Timestamp should display below or next to received message.", "Medium / Medium"],
    
    # Verify Enter key to send - Positive
    ["MSG_028", "Positive", "Verify Enter key to send", "User is logged in and message typed", "Verify Enter key sends message", "1. Type a message.\n2. Press Enter key.\n3. Observe chat.", "Message should be sent when Enter key is pressed.", "High / High"],
    ["MSG_029", "Positive", "Verify Enter key to send", "User is logged in and message typed", "Verify Shift+Enter creates new line", "1. Type a message.\n2. Press Shift+Enter.\n3. Continue typing.", "New line should be created; message should not be sent.", "Medium / Medium"],
    # Verify Enter key to send - Negative
    ["MSG_030", "Negative", "Verify Enter key to send", "User is logged in and input is empty", "Verify Enter key with empty input", "1. Leave input field empty.\n2. Press Enter key.", "Nothing should happen; empty message should not be sent.", "Medium / Medium"],
    
    # Verify Message input clear after send - Positive
    ["MSG_031", "Positive", "Verify Message input clear after send", "User has sent a message", "Verify input field clears after sending", "1. Type a message.\n2. Click send.\n3. Observe input field.", "Input field should be cleared after message is sent successfully.", "High / High"],
    
    # Verify Real-time message delivery - Positive
    ["MSG_032", "Positive", "Verify Real-time message delivery", "Two users are in same chat", "Verify message appears instantly for recipient", "1. User A sends message.\n2. User B observes chat.", "Message should appear in User B's chat within seconds without refresh.", "High / High"],
    ["MSG_033", "Positive", "Verify Real-time message delivery", "Two users are in same chat", "Verify typing indicator", "1. User A starts typing.\n2. User B observes chat.", "Typing indicator should appear for User B (e.g., \"User A is typing...\").", "Medium / Medium"],
    
    # Verify Message scroll behavior - Positive
    ["MSG_034", "Positive", "Verify Message scroll behavior", "User is in chat with many messages", "Verify auto-scroll to new message", "1. Be at bottom of chat.\n2. Send or receive new message.", "Chat should auto-scroll to show the new message.", "High / Medium"],
    ["MSG_035", "Positive", "Verify Message scroll behavior", "User is in chat with many messages", "Verify scroll up to view history", "1. Open chat with many messages.\n2. Scroll up.", "Older messages should load; scrolling should be smooth.", "Medium / Medium"],
    ["MSG_036", "Positive", "Verify Message scroll behavior", "User is scrolled up in chat", "Verify new message notification when scrolled up", "1. Scroll up in chat.\n2. Receive new message.", "Notification or \"New message\" button should appear to scroll to bottom.", "Medium / Medium"],
    
    # Verify Attachment button - Positive
    ["MSG_037", "Positive", "Verify Attachment button", "User is logged in and chat is open", "Verify attachment button is visible", "1. Open any chat conversation.\n2. Observe attachment button.", "Attachment button (paperclip/+ icon) should be visible near input field.", "High / High"],
    ["MSG_038", "Positive", "Verify Attachment button", "User is logged in and chat is open", "Verify attachment button click opens options", "1. Open any chat conversation.\n2. Click attachment button.", "Attachment options should appear (image, file, camera, etc.).", "High / High"],
    
    # Verify Emoji button - Positive
    ["MSG_039", "Positive", "Verify Emoji button", "User is logged in and chat is open", "Verify emoji button is visible", "1. Open any chat conversation.\n2. Observe emoji button.", "Emoji button (smiley icon) should be visible near input field.", "Medium / Medium"],
    ["MSG_040", "Positive", "Verify Emoji button", "User is logged in and chat is open", "Verify emoji button click opens picker", "1. Open any chat conversation.\n2. Click emoji button.", "Emoji picker should open with emoji categories.", "Medium / Medium"],
    ["MSG_041", "Positive", "Verify Emoji button", "User is logged in and emoji picker open", "Verify selecting emoji adds to input", "1. Open emoji picker.\n2. Click on an emoji.", "Selected emoji should be added to message input field.", "Medium / Medium"],
    
    # Verify UI responsiveness
    ["MSG_042", "Positive", "Verify UI responsiveness", "User is logged in", "Verify message input on desktop", "1. Open application on desktop.\n2. Open chat.\n3. Resize browser.", "Message input and send button should remain functional and visible.", "High / Medium"],
    ["MSG_043", "Positive", "Verify UI responsiveness", "User is logged in", "Verify message input on mobile", "1. Open application on mobile.\n2. Open chat.\n3. Check keyboard interaction.", "Input field should work with mobile keyboard; send button should be accessible.", "High / Medium"],
    
    # Verify Message loading - Negative
    ["MSG_044", "Negative", "Verify Message loading", "User is logged in, slow network", "Verify message sending with slow network", "1. Simulate slow network.\n2. Send a message.\n3. Observe behavior.", "Loading indicator should show; message should eventually send or show retry option.", "Medium / Medium"],
    ["MSG_045", "Negative", "Verify Message loading", "User is logged in, network error", "Verify message retry on failure", "1. Send message during network error.\n2. Observe failed message.\n3. Click retry.", "Retry option should be available; message should resend on retry.", "High / High"],
    
    # Verify Attachment feature - Positive
    ["MSG_046", "Positive", "Verify Attachment feature", "User is logged in and chat is open", "Verify attachment icon is visible", "1. Open any chat conversation.\n2. Observe attachment icon near input field.", "Attachment icon (paperclip/+ icon) should be visible.", "High / High"],
    ["MSG_047", "Positive", "Verify Attachment feature", "User is logged in and chat is open", "Verify attachment options menu", "1. Click on attachment icon.\n2. Observe options menu.", "Options menu should display (Image, Video, Document, Camera, etc.).", "High / High"],
    ["MSG_048", "Positive", "Verify Attachment feature", "User is logged in and attachment menu open", "Verify sending image attachment", "1. Click attachment icon.\n2. Select Image option.\n3. Choose an image.\n4. Send.", "Image should be uploaded and sent; preview should display in chat.", "High / High"],
    ["MSG_049", "Positive", "Verify Attachment feature", "User is logged in and attachment menu open", "Verify sending document attachment", "1. Click attachment icon.\n2. Select Document option.\n3. Choose a file.\n4. Send.", "Document should be uploaded and sent; file icon with name should display.", "High / High"],
    ["MSG_050", "Positive", "Verify Attachment feature", "User is logged in and attachment menu open", "Verify sending video attachment", "1. Click attachment icon.\n2. Select Video option.\n3. Choose a video.\n4. Send.", "Video should be uploaded and sent; video thumbnail should display.", "Medium / Medium"],
    ["MSG_051", "Positive", "Verify Attachment feature", "User is logged in and attachment menu open", "Verify attachment upload progress", "1. Click attachment icon.\n2. Select a large file.\n3. Observe upload.", "Upload progress indicator should display during file upload.", "Medium / Medium"],
    # Verify Attachment feature - Negative
    ["MSG_052", "Negative", "Verify Attachment feature", "User is logged in and chat is open", "Verify unsupported file type handling", "1. Click attachment icon.\n2. Try to select unsupported file type.", "Error message should display or file should not be selectable.", "Medium / Medium"],
    ["MSG_053", "Negative", "Verify Attachment feature", "User is logged in and chat is open", "Verify file size limit handling", "1. Click attachment icon.\n2. Select file exceeding size limit.", "Error message should display indicating file size limit exceeded.", "High / Medium"],
    ["MSG_054", "Negative", "Verify Attachment feature", "User is logged in, network error", "Verify attachment upload failure handling", "1. Disconnect network.\n2. Try to send attachment.", "Error message should display; retry option should be available.", "High / High"],
    
    # Verify Voice Recording feature - Positive
    ["MSG_055", "Positive", "Verify Voice Recording feature", "User is logged in and chat is open", "Verify recording button is visible", "1. Open any chat conversation.\n2. Observe microphone/recording icon.", "Recording button (microphone icon) should be visible near input field.", "High / High"],
    ["MSG_056", "Positive", "Verify Voice Recording feature", "User is logged in and chat is open", "Verify recording starts on button press", "1. Press and hold recording button.\n2. Observe recording indicator.", "Recording should start; timer and waveform indicator should display.", "High / High"],
    ["MSG_057", "Positive", "Verify Voice Recording feature", "User is recording voice message", "Verify recording timer display", "1. Start recording.\n2. Observe timer.", "Recording duration timer should display and increment.", "Medium / Medium"],
    ["MSG_058", "Positive", "Verify Voice Recording feature", "User is recording voice message", "Verify sending voice message", "1. Record a voice message.\n2. Release button or click send.\n3. Observe chat.", "Voice message should be sent and appear in chat with play button.", "High / High"],
    ["MSG_059", "Positive", "Verify Voice Recording feature", "User is recording voice message", "Verify cancel recording", "1. Start recording.\n2. Slide to cancel or click cancel button.", "Recording should be cancelled; no message should be sent.", "High / Medium"],
    ["MSG_060", "Positive", "Verify Voice Recording feature", "User has received voice message", "Verify playing received voice message", "1. Receive a voice message.\n2. Click play button.", "Voice message should play; progress indicator should show.", "High / High"],
    # Verify Voice Recording feature - Negative
    ["MSG_061", "Negative", "Verify Voice Recording feature", "User is logged in, microphone permission denied", "Verify recording without microphone permission", "1. Deny microphone permission.\n2. Try to record voice message.", "Permission request should appear or error message should display.", "High / High"],
    ["MSG_062", "Negative", "Verify Voice Recording feature", "User is logged in and chat is open", "Verify very short recording handling", "1. Press and release recording button quickly (< 1 second).", "Recording should be cancelled or warning should display.", "Low / Low"],
    
    # Verify Emoji feature - Positive
    ["MSG_063", "Positive", "Verify Emoji feature", "User is logged in and chat is open", "Verify emoji button is visible", "1. Open any chat conversation.\n2. Observe emoji button.", "Emoji button (smiley face icon) should be visible near input field.", "High / High"],
    ["MSG_064", "Positive", "Verify Emoji feature", "User is logged in and chat is open", "Verify emoji picker opens", "1. Click on emoji button.\n2. Observe emoji picker.", "Emoji picker should open with categories (Smileys, Animals, Food, etc.).", "High / High"],
    ["MSG_065", "Positive", "Verify Emoji feature", "User is logged in and emoji picker open", "Verify emoji categories navigation", "1. Open emoji picker.\n2. Click on different category tabs.", "Different emoji categories should display when tabs are clicked.", "Medium / Medium"],
    ["MSG_066", "Positive", "Verify Emoji feature", "User is logged in and emoji picker open", "Verify selecting emoji adds to input", "1. Open emoji picker.\n2. Click on any emoji.", "Selected emoji should be added to message input field at cursor position.", "High / High"],
    ["MSG_067", "Positive", "Verify Emoji feature", "User is logged in and emoji picker open", "Verify multiple emoji selection", "1. Open emoji picker.\n2. Click multiple emojis.", "All selected emojis should be added to input field.", "Medium / Medium"],
    ["MSG_068", "Positive", "Verify Emoji feature", "User is logged in and emoji picker open", "Verify emoji search functionality", "1. Open emoji picker.\n2. Type in emoji search field.\n3. Observe results.", "Matching emojis should display based on search term.", "Medium / Medium"],
    ["MSG_069", "Positive", "Verify Emoji feature", "User is logged in and emoji picker open", "Verify recent emojis section", "1. Send message with emoji.\n2. Open emoji picker again.\n3. Check recent section.", "Recently used emojis should appear in recent/frequently used section.", "Low / Low"],
    ["MSG_070", "Positive", "Verify Emoji feature", "User is logged in and emoji picker open", "Verify closing emoji picker", "1. Open emoji picker.\n2. Click outside picker or close button.", "Emoji picker should close.", "Medium / Medium"],
    
    # Verify Sticker feature - Positive
    ["MSG_071", "Positive", "Verify Sticker feature", "User is logged in and chat is open", "Verify sticker button/tab is visible", "1. Open any chat conversation.\n2. Observe sticker option.", "Sticker button or tab should be visible (may be in emoji picker).", "High / High"],
    ["MSG_072", "Positive", "Verify Sticker feature", "User is logged in and chat is open", "Verify sticker picker opens", "1. Click on sticker button/tab.\n2. Observe sticker picker.", "Sticker picker should open with sticker packs.", "High / High"],
    ["MSG_073", "Positive", "Verify Sticker feature", "User is logged in and sticker picker open", "Verify sticker packs display", "1. Open sticker picker.\n2. Observe available sticker packs.", "Different sticker packs should be visible and selectable.", "Medium / Medium"],
    ["MSG_074", "Positive", "Verify Sticker feature", "User is logged in and sticker picker open", "Verify sending sticker", "1. Open sticker picker.\n2. Click on any sticker.", "Sticker should be sent immediately and appear in chat.", "High / High"],
    ["MSG_075", "Positive", "Verify Sticker feature", "User has received sticker", "Verify received sticker display", "1. Receive a sticker from another user.\n2. Observe chat.", "Sticker should display properly in chat on left side.", "Medium / Medium"],
    ["MSG_076", "Positive", "Verify Sticker feature", "User is logged in and sticker picker open", "Verify sticker pack switching", "1. Open sticker picker.\n2. Click on different sticker pack.", "Selected sticker pack's stickers should display.", "Medium / Medium"],
    # Verify Sticker feature - Negative
    ["MSG_077", "Negative", "Verify Sticker feature", "User is logged in, no sticker packs available", "Verify empty sticker state", "1. Open sticker picker with no packs.\n2. Observe display.", "Empty state or option to download sticker packs should display.", "Low / Low"],
    
    # Verify @all mention feature - Positive
    ["MSG_078", "Positive", "Verify @all mention feature", "User is logged in and group chat is open", "Verify typing @all shows suggestion", "1. Open group chat.\n2. Type \"@all\" in input field.", "\"@all\" suggestion should appear to mention all group members.", "High / High"],
    ["MSG_079", "Positive", "Verify @all mention feature", "User is logged in and group chat is open", "Verify selecting @all mention", "1. Type \"@all\".\n2. Select @all from suggestions.", "\"@all\" should be added to message with special formatting.", "High / High"],
    ["MSG_080", "Positive", "Verify @all mention feature", "User is logged in and group chat is open", "Verify sending message with @all", "1. Type message with @all.\n2. Send message.", "Message should be sent; all group members should receive notification.", "High / High"],
    ["MSG_081", "Positive", "Verify @all mention feature", "User has received @all mention", "Verify @all mention notification", "1. Another user sends @all in group.\n2. Observe notification.", "User should receive notification for @all mention.", "High / High"],
    ["MSG_082", "Positive", "Verify @all mention feature", "User has received @all mention", "Verify @all highlight in message", "1. Receive message with @all.\n2. Observe message display.", "\"@all\" should be highlighted/styled differently in the message.", "Medium / Medium"],
    # Verify @all mention feature - Negative
    ["MSG_083", "Negative", "Verify @all mention feature", "User is logged in and one-on-one chat is open", "Verify @all not available in direct chat", "1. Open one-on-one chat.\n2. Type \"@all\".", "@all suggestion should not appear in direct/one-on-one chats.", "Medium / Medium"],
    
    # Verify @ mention feature - Positive
    ["MSG_084", "Positive", "Verify @ mention feature", "User is logged in and group chat is open", "Verify typing @ shows member suggestions", "1. Open group chat.\n2. Type \"@\" in input field.", "List of group members should appear as suggestions.", "High / High"],
    ["MSG_085", "Positive", "Verify @ mention feature", "User is logged in and group chat is open", "Verify filtering members by name", "1. Type \"@Jo\" in input field.\n2. Observe suggestions.", "Only members with names starting with \"Jo\" should appear (e.g., John).", "High / High"],
    ["MSG_086", "Positive", "Verify @ mention feature", "User is logged in and group chat is open", "Verify selecting member from suggestions", "1. Type \"@\".\n2. Click on a member name.", "Selected member's name should be added to message with @ prefix.", "High / High"],
    ["MSG_087", "Positive", "Verify @ mention feature", "User is logged in and group chat is open", "Verify sending message with @ mention", "1. Type message with @username.\n2. Send message.", "Message should be sent; mentioned user should receive notification.", "High / High"],
    ["MSG_088", "Positive", "Verify @ mention feature", "User has been mentioned", "Verify @ mention notification", "1. Another user mentions you with @.\n2. Observe notification.", "User should receive notification for being mentioned.", "High / High"],
    ["MSG_089", "Positive", "Verify @ mention feature", "User has been mentioned", "Verify @ mention highlight in message", "1. Receive message where you are mentioned.\n2. Observe message.", "Your @mention should be highlighted/styled differently.", "Medium / Medium"],
    ["MSG_090", "Positive", "Verify @ mention feature", "User is logged in and group chat is open", "Verify multiple @ mentions in one message", "1. Type message with multiple @mentions.\n2. Send message.", "All mentioned users should receive notifications.", "Medium / Medium"],
    ["MSG_091", "Positive", "Verify @ mention feature", "User is logged in and group chat is open", "Verify @ mention with profile picture in suggestions", "1. Type \"@\" in input field.\n2. Observe suggestion list.", "Member suggestions should show profile pictures alongside names.", "Low / Low"],
    # Verify @ mention feature - Negative
    ["MSG_092", "Negative", "Verify @ mention feature", "User is logged in and group chat is open", "Verify @ with no matching members", "1. Type \"@xyz123\" in input field.\n2. Observe suggestions.", "No suggestions should appear or \"No members found\" message.", "Medium / Medium"],
    ["MSG_093", "Negative", "Verify @ mention feature", "User is logged in and one-on-one chat is open", "Verify @ mention in direct chat", "1. Open one-on-one chat.\n2. Type \"@\".", "Only the other user should appear in suggestions (or feature disabled).", "Low / Low"],
]

# Styles
header_font = Font(bold=True)
header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Start from row 1
start_row = 1

# Write headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=start_row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# First pass: Write all data
for row_idx, row_data in enumerate(data, start_row + 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = thin_border

# Second pass: Merge cells
def merge_similar_cells(ws, col_idx, start_row, data):
    data_start = start_row + 1
    i = 0
    while i < len(data):
        current_sentiment = data[i][1]
        current_scenario = data[i][2]
        current_precondition = data[i][3]
        merge_start = data_start + i
        merge_end = merge_start
        
        j = i + 1
        while j < len(data):
            if data[j][1] == current_sentiment and data[j][2] == current_scenario:
                merge_end = data_start + j
                j += 1
            else:
                break
        
        if merge_end > merge_start:
            ws.merge_cells(start_row=merge_start, start_column=2, end_row=merge_end, end_column=2)
            ws.cell(row=merge_start, column=2).alignment = Alignment(vertical='top', horizontal='center', wrap_text=True)
        
        if merge_end > merge_start:
            ws.merge_cells(start_row=merge_start, start_column=3, end_row=merge_end, end_column=3)
            ws.cell(row=merge_start, column=3).alignment = Alignment(vertical='top', wrap_text=True)
        
        k = i
        while k <= j - 1:
            precond_start = data_start + k
            precond_end = precond_start
            current_precond = data[k][3]
            
            m = k + 1
            while m < j:
                if data[m][3] == current_precond:
                    precond_end = data_start + m
                    m += 1
                else:
                    break
            
            if precond_end > precond_start:
                ws.merge_cells(start_row=precond_start, start_column=4, end_row=precond_end, end_column=4)
                ws.cell(row=precond_start, column=4).alignment = Alignment(vertical='top', wrap_text=True)
            
            k = m
        
        i = j

merge_similar_cells(ws, 2, start_row, data)

# Set column widths
column_widths = [10, 10, 26, 28, 32, 36, 42, 18]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

ws.row_dimensions[start_row].height = 20
ws.freeze_panes = 'A2'

wb.save("Send_Message/Send_Message_Test_Cases.xlsx")
print("Send Message Module Excel file created with 93 test cases!")
