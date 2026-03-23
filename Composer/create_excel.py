from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Composer Test Cases"

# Headers
headers = ["TestCase_ID", "Sentiment", "Test Scenario", "Precondition", "Test Case", "Steps", "Expected Results", "Priority / Severity"]

# Data for Composer module - Advanced level test cases
data = [
    # Verify Send Button - Positive
    ["COMP_001", "Positive", "Verify Send Button", "User is logged in and chat is open", "Verify send button sends text message", "1. Type message.\n2. Click send button.", "Message should be sent and appear in chat with timestamp.", "High / High"],
    ["COMP_002", "Positive", "Verify Send Button", "User is logged in and chat is open", "Verify Enter key sends message", "1. Type message.\n2. Press Enter key.", "Message should be sent.", "High / High"],
    # Verify Send Button - Negative
    ["COMP_003", "Negative", "Verify Send Button", "User is logged in and chat is open", "Verify empty message not sent", "1. Leave input empty.\n2. Click send.", "Message should not be sent.", "High / High"],
    ["COMP_004", "Negative", "Verify Send Button", "User is logged in, network disconnected", "Verify send fails without network", "1. Disconnect network.\n2. Send message.", "Error should display with retry option.", "High / High"],
    
    # Verify Attach Button - Positive
    ["COMP_005", "Positive", "Verify Attach Button", "User is logged in and chat is open", "Verify attach menu opens", "1. Click attach (+) button.", "Attachment options should display (Image, Video, Document).", "High / High"],
    ["COMP_006", "Positive", "Verify Attach Button", "User is logged in and chat is open", "Verify sending image attachment", "1. Click attach.\n2. Select image.\n3. Send.", "Image should upload and display in chat.", "High / High"],
    ["COMP_007", "Positive", "Verify Attach Button", "User is logged in and chat is open", "Verify sending document attachment", "1. Click attach.\n2. Select document.\n3. Send.", "Document should upload with file icon and name.", "High / High"],
    # Verify Attach Button - Negative
    ["COMP_008", "Negative", "Verify Attach Button", "User is logged in and chat is open", "Verify file size limit exceeded", "1. Select file exceeding size limit.", "Error message should display.", "High / Medium"],
    ["COMP_009", "Negative", "Verify Attach Button", "User is logged in, network error", "Verify attachment upload failure", "1. Disconnect network.\n2. Send attachment.", "Error should display with retry option.", "High / High"],
    
    # Verify Recording Button - Positive
    ["COMP_010", "Positive", "Verify Recording Button", "User is logged in and chat is open", "Verify voice recording and send", "1. Press and hold mic button.\n2. Record.\n3. Release to send.", "Voice message should be sent with play button.", "High / High"],
    ["COMP_011", "Positive", "Verify Recording Button", "User is logged in and chat is open", "Verify cancel recording", "1. Start recording.\n2. Slide to cancel.", "Recording should be cancelled.", "High / Medium"],
    ["COMP_012", "Positive", "Verify Recording Button", "User has received voice message", "Verify playing voice message", "1. Click play on received voice message.", "Voice message should play with progress.", "High / High"],
    # Verify Recording Button - Negative
    ["COMP_013", "Negative", "Verify Recording Button", "User is logged in, mic permission denied", "Verify recording without permission", "1. Deny mic permission.\n2. Try to record.", "Permission request or error should display.", "High / High"],
    
    # Verify Emoji Button - Positive
    ["COMP_014", "Positive", "Verify Emoji Button", "User is logged in and chat is open", "Verify emoji picker opens and selection", "1. Click emoji button.\n2. Select emoji.", "Emoji should be added to input field.", "High / High"],
    ["COMP_015", "Positive", "Verify Emoji Button", "User is logged in and chat is open", "Verify sending message with emoji", "1. Add emoji to message.\n2. Send.", "Message with emoji should display correctly.", "High / High"],
    
    # Verify Sticker Button - Positive
    ["COMP_016", "Positive", "Verify Sticker Button", "User is logged in and chat is open", "Verify sticker picker opens", "1. Click sticker button.", "Sticker picker should open with sticker packs.", "High / High"],
    ["COMP_017", "Positive", "Verify Sticker Button", "User is logged in and chat is open", "Verify sending sticker", "1. Open sticker picker.\n2. Click sticker.", "Sticker should be sent and display in chat.", "High / High"],
    
    # Verify @all mention - Positive
    ["COMP_018", "Positive", "Verify @all mention", "User is logged in and group chat is open", "Verify @all mention in group", "1. Type \"@all\".\n2. Select from suggestions.\n3. Send.", "Message should be sent; all members notified.", "High / High"],
    # Verify @all mention - Negative
    ["COMP_019", "Negative", "Verify @all mention", "User is logged in and one-on-one chat is open", "Verify @all not available in direct chat", "1. Open direct chat.\n2. Type \"@all\".", "@all suggestion should not appear.", "Medium / Medium"],
    
    # Verify @ mention - Positive
    ["COMP_020", "Positive", "Verify @ mention", "User is logged in and group chat is open", "Verify @ shows member suggestions", "1. Type \"@\" in group chat.", "Member list with avatars should appear.", "High / High"],
    ["COMP_021", "Positive", "Verify @ mention", "User is logged in and group chat is open", "Verify filtering members by name", "1. Type \"@Geo\".", "Only matching members should appear.", "High / High"],
    ["COMP_022", "Positive", "Verify @ mention", "User is logged in and group chat is open", "Verify sending @ mention message", "1. Select member.\n2. Send message.", "Mentioned user should receive notification.", "High / High"],
    # Verify @ mention - Negative
    ["COMP_023", "Negative", "Verify @ mention", "User is logged in and group chat is open", "Verify @ with no matching members", "1. Type \"@xyz123\".", "No suggestions should appear.", "Medium / Medium"],
    
    # Verify Composer in One-on-One vs Group - Positive
    ["COMP_024", "Positive", "Verify Composer in One-on-One chat", "User is logged in and one-on-one chat is open", "Verify all features work in direct chat", "1. Test send, attach, record, emoji, sticker.", "All features should work except @all.", "High / High"],
    ["COMP_025", "Positive", "Verify Composer in Group chat", "User is logged in and group chat is open", "Verify all features work in group chat", "1. Test all features including @mentions.", "All features should work including @all and @.", "High / High"],
    
    # Verify UI responsiveness - Positive
    ["COMP_026", "Positive", "Verify UI responsiveness", "User is logged in", "Verify composer on desktop and mobile", "1. Test on desktop.\n2. Test on mobile.", "Composer should be functional on both platforms.", "High / Medium"],
]

# Styles
header_font = Font(bold=True)
header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Start from row 1
start_row = 1

# Write headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=start_row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# First pass: Write all data
for row_idx, row_data in enumerate(data, start_row + 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = thin_border

# Second pass: Merge cells
def merge_similar_cells(ws, col_idx, start_row, data):
    data_start = start_row + 1
    i = 0
    while i < len(data):
        current_sentiment = data[i][1]
        current_scenario = data[i][2]
        merge_start = data_start + i
        merge_end = merge_start
        
        j = i + 1
        while j < len(data):
            if data[j][1] == current_sentiment and data[j][2] == current_scenario:
                merge_end = data_start + j
                j += 1
            else:
                break
        
        if merge_end > merge_start:
            ws.merge_cells(start_row=merge_start, start_column=2, end_row=merge_end, end_column=2)
            ws.cell(row=merge_start, column=2).alignment = Alignment(vertical='top', horizontal='center', wrap_text=True)
        
        if merge_end > merge_start:
            ws.merge_cells(start_row=merge_start, start_column=3, end_row=merge_end, end_column=3)
            ws.cell(row=merge_start, column=3).alignment = Alignment(vertical='top', wrap_text=True)
        
        k = i
        while k <= j - 1:
            precond_start = data_start + k
            precond_end = precond_start
            current_precond = data[k][3]
            
            m = k + 1
            while m < j:
                if data[m][3] == current_precond:
                    precond_end = data_start + m
                    m += 1
                else:
                    break
            
            if precond_end > precond_start:
                ws.merge_cells(start_row=precond_start, start_column=4, end_row=precond_end, end_column=4)
                ws.cell(row=precond_start, column=4).alignment = Alignment(vertical='top', wrap_text=True)
            
            k = m
        
        i = j

merge_similar_cells(ws, 2, start_row, data)

# Set column widths
column_widths = [10, 10, 26, 32, 34, 38, 44, 18]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

ws.row_dimensions[start_row].height = 20
ws.freeze_panes = 'A2'

wb.save("Composer/Composer_Test_Cases.xlsx")
print("Composer Module Excel file created with 26 advanced test cases!")
