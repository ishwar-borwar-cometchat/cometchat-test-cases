from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Conversation List Test Cases"

# Headers
headers = ["TestCase_ID", "Sentiment", "Test Scenario", "Precondition", "Test Case", "Steps", "Expected Results", "Priority / Severity"]

# Data for Conversation List module - Advanced level test cases
data = [
    # Verify User and Group Listing - Positive
    ["CL_001", "Positive", "Verify User and Group Listing", "User is logged in and has conversations", "Verify conversation list displays users and groups", "1. Navigate to Chats tab.\n2. Observe conversation list.", "Both user chats and group chats should display in the list.", "High / High"],
    ["CL_002", "Positive", "Verify User and Group Listing", "User is logged in and has conversations", "Verify user conversation displays name and avatar", "1. Observe user conversation in list.", "User name, profile picture, and last message should display.", "High / High"],
    ["CL_003", "Positive", "Verify User and Group Listing", "User is logged in and has conversations", "Verify group conversation displays name and avatar", "1. Observe group conversation in list.", "Group name, group avatar/initials, and last message should display.", "High / High"],
    ["CL_004", "Positive", "Verify User and Group Listing", "User is logged in and has conversations", "Verify online status indicator", "1. Observe avatar in conversation list.", "Online users/groups should show green dot indicator.", "Medium / Medium"],
    ["CL_005", "Positive", "Verify User and Group Listing", "User is logged in and has conversations", "Verify conversations sorted by recent activity", "1. Send message in older conversation.\n2. Observe list order.", "Most recent conversation should move to top.", "High / High"],
    ["CL_006", "Positive", "Verify User and Group Listing", "User is logged in and has conversations", "Verify clicking conversation opens chat", "1. Click on any conversation.", "Chat window should open with message history.", "High / High"],
    # Verify User and Group Listing - Negative
    ["CL_007", "Negative", "Verify User and Group Listing", "User is logged in with no conversations", "Verify empty conversation list state", "1. Login with new account.\n2. Navigate to Chats.", "Empty state message should display.", "Medium / Medium"],
    
    # Verify Delete Conversation - Positive
    ["CL_008", "Positive", "Verify Delete Conversation", "User is logged in and has conversations", "Verify delete icon appears on swipe/hover", "1. Swipe left on conversation (mobile) or hover (desktop).", "Delete icon (trash) should appear.", "High / High"],
    ["CL_009", "Positive", "Verify Delete Conversation", "User is logged in and has conversations", "Verify delete confirmation dialog", "1. Click delete icon on conversation.", "Confirmation dialog should appear.", "High / High"],
    ["CL_010", "Positive", "Verify Delete Conversation", "User is logged in and delete dialog shown", "Verify conversation deleted on confirm", "1. Click confirm in delete dialog.", "Conversation should be removed from list.", "High / High"],
    ["CL_011", "Positive", "Verify Delete Conversation", "User is logged in and delete dialog shown", "Verify cancel delete keeps conversation", "1. Click cancel in delete dialog.", "Conversation should remain in list.", "Medium / Medium"],
    # Verify Delete Conversation - Negative
    ["CL_012", "Negative", "Verify Delete Conversation", "User is logged in, network disconnected", "Verify delete fails without network", "1. Disconnect network.\n2. Try to delete conversation.", "Error message should display.", "High / High"],
    
    # Verify Check Day (Date/Time display) - Positive
    ["CL_013", "Positive", "Verify Check Day", "User is logged in and has conversations", "Verify today's conversation shows time", "1. Observe conversation from today.", "Time should display (e.g., \"02:35 AM\").", "High / High"],
    ["CL_014", "Positive", "Verify Check Day", "User is logged in and has conversations", "Verify yesterday's conversation shows 'Yesterday'", "1. Observe conversation from yesterday.", "\"Yesterday\" should display instead of date.", "High / High"],
    ["CL_015", "Positive", "Verify Check Day", "User is logged in and has conversations", "Verify older conversation shows date", "1. Observe conversation older than yesterday.", "Date should display (e.g., \"05/02/2026\").", "High / High"],
    ["CL_016", "Positive", "Verify Check Day", "User is logged in and has conversations", "Verify date updates after midnight", "1. Check conversation at 11:59 PM.\n2. Check again after midnight.", "Date/time label should update appropriately.", "Medium / Medium"],
    
    # Verify Preview Message - Positive
    ["CL_017", "Positive", "Verify Preview Message", "User is logged in and has conversations", "Verify last message preview displays", "1. Observe conversation in list.", "Last message preview should display below name.", "High / High"],
    ["CL_018", "Positive", "Verify Preview Message", "User is logged in and has conversations", "Verify 'You:' prefix for sent messages", "1. Send message.\n2. Observe conversation preview.", "Preview should show \"You: [message]\" for sent messages.", "High / High"],
    ["CL_019", "Positive", "Verify Preview Message", "User is logged in and has conversations", "Verify message status checkmark", "1. Observe sent message preview.", "Checkmark should display (single=sent, double=delivered).", "High / High"],
    ["CL_020", "Positive", "Verify Preview Message", "User is logged in and has conversations", "Verify long message truncation", "1. Send long message.\n2. Observe preview.", "Long message should be truncated with ellipsis.", "Medium / Medium"],
    ["CL_021", "Positive", "Verify Preview Message", "User is logged in and has conversations", "Verify image/attachment preview", "1. Send image.\n2. Observe preview.", "Preview should show attachment indicator (e.g., \"📷 Photo\").", "Medium / Medium"],
    ["CL_022", "Positive", "Verify Preview Message", "User is logged in and has conversations", "Verify preview updates on new message", "1. Receive new message.\n2. Observe preview.", "Preview should update to show new message.", "High / High"],
    # Verify Preview Message - Negative
    ["CL_023", "Negative", "Verify Preview Message", "User is logged in with new conversation", "Verify empty preview for new conversation", "1. Start new conversation without messages.", "No preview or placeholder should display.", "Low / Low"],
    
    # Verify UI responsiveness - Positive
    ["CL_024", "Positive", "Verify UI responsiveness", "User is logged in", "Verify conversation list on desktop and mobile", "1. Test on desktop.\n2. Test on mobile.", "List should be functional on both platforms.", "High / Medium"],
]

# Styles
header_font = Font(bold=True)
header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Start from row 1
start_row = 1

# Write headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=start_row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# First pass: Write all data
for row_idx, row_data in enumerate(data, start_row + 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = thin_border

# Second pass: Merge cells
def merge_similar_cells(ws, col_idx, start_row, data):
    data_start = start_row + 1
    i = 0
    while i < len(data):
        current_sentiment = data[i][1]
        current_scenario = data[i][2]
        merge_start = data_start + i
        merge_end = merge_start
        
        j = i + 1
        while j < len(data):
            if data[j][1] == current_sentiment and data[j][2] == current_scenario:
                merge_end = data_start + j
                j += 1
            else:
                break
        
        if merge_end > merge_start:
            ws.merge_cells(start_row=merge_start, start_column=2, end_row=merge_end, end_column=2)
            ws.cell(row=merge_start, column=2).alignment = Alignment(vertical='top', horizontal='center', wrap_text=True)
        
        if merge_end > merge_start:
            ws.merge_cells(start_row=merge_start, start_column=3, end_row=merge_end, end_column=3)
            ws.cell(row=merge_start, column=3).alignment = Alignment(vertical='top', wrap_text=True)
        
        k = i
        while k <= j - 1:
            precond_start = data_start + k
            precond_end = precond_start
            current_precond = data[k][3]
            
            m = k + 1
            while m < j:
                if data[m][3] == current_precond:
                    precond_end = data_start + m
                    m += 1
                else:
                    break
            
            if precond_end > precond_start:
                ws.merge_cells(start_row=precond_start, start_column=4, end_row=precond_end, end_column=4)
                ws.cell(row=precond_start, column=4).alignment = Alignment(vertical='top', wrap_text=True)
            
            k = m
        
        i = j

merge_similar_cells(ws, 2, start_row, data)

# Set column widths
column_widths = [10, 10, 26, 32, 34, 38, 44, 18]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

ws.row_dimensions[start_row].height = 20
ws.freeze_panes = 'A2'

wb.save("Conversation_List/Conversation_List_Test_Cases.xlsx")
print("Conversation List Module Excel file created with 24 test cases!")
