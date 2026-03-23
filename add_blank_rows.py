import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def style_sheet(ws):
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap = Alignment(wrap_text=True, vertical='top')

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        cell.border = thin_border

    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = wrap
            cell.border = thin_border

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 45
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 18

SHEET_ORDER = ['Functional', 'Integration', 'Regression', 'Security', 'Edge Cases', 'Boundary Values', 'Equivalence Partitioning']

TAB_COLORS = {
    'Functional': '4472C4',
    'Integration': '70AD47',
    'Regression': 'FFC000',
    'Security': 'FF0000',
    'Edge Cases': 'ED7D31',
    'Boundary Values': '7030A0',
    'Equivalence Partitioning': '00B0F0',
}

files = [
    ('Conversation_List/Conversation_List_Test_Cases.xlsx', 'CL'),
    ('Message_List/Message_List_Test_Cases.xlsx', 'ML'),
    ('Advanced_Search/Advanced_Search_Test_Cases.xlsx', 'AS'),
    ('Conversation_Search/Conversation_Search_Test_Cases.xlsx', 'CS'),
    ('Composer/Composer_Test_Cases.xlsx', 'COMP'),
    ('Send_Message/Send_Message_Test_Cases.xlsx', 'MSG'),
    ('Send_Threaded_Message/Send_Threaded_Message_Test_Cases.xlsx', 'STM'),
    ('Group_Actions/Group_Actions_Test_Cases.xlsx', 'GA'),
    ('Groups_Module/Groups_Module_Test_Cases.xlsx', 'GROUPS'),
    ('User_Module/Users_Module_Test_Script.xlsx', 'USERS'),
    ('Call_Module/Calls_Module_Test_Cases.xlsx', 'CALLS'),
]

def classify_test_case(scenario_text):
    s = (scenario_text or '').lower()
    if 'integration' in s:
        return 'Integration'
    elif 'regression' in s:
        return 'Regression'
    elif 'security' in s:
        return 'Security'
    elif 'edge case' in s:
        return 'Edge Cases'
    elif 'boundary' in s:
        return 'Boundary Values'
    elif 'equivalence' in s:
        return 'Equivalence Partitioning'
    else:
        return 'Functional'


for filepath, prefix in files:
    full_path = 'cometchat-test-cases/' + filepath
    wb = openpyxl.load_workbook(full_path)

    # Read headers from first sheet
    first_ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in first_ws[1]]
    num_cols = len(headers)

    # Read ALL data from ALL sheets, classify, and track sentiment
    classified = {t: [] for t in SHEET_ORDER}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        current_scenario = ''
        current_sentiment = ''
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if all(v is None for v in row):
                continue
            scenario = row[2]
            sentiment = row[1]
            if scenario:
                current_scenario = scenario
            if sentiment:
                current_sentiment = sentiment
            test_type = classify_test_case(current_scenario)
            classified[test_type].append((current_sentiment, list(row)))

    # Create new workbook
    new_wb = openpyxl.Workbook()
    new_wb.remove(new_wb.active)

    for sheet_type in SHEET_ORDER:
        rows_with_sentiment = classified[sheet_type]
        if not rows_with_sentiment:
            continue

        sheet_name = sheet_type[:31]
        new_ws = new_wb.create_sheet(title=sheet_name)
        new_ws.sheet_properties.tabColor = TAB_COLORS[sheet_type]

        # Write headers
        new_ws.append(headers)

        # First section: ALL Positive, then blank row, then ALL Negative
        positive_rows = [r for s, r in rows_with_sentiment if s == 'Positive']
        negative_rows = [r for s, r in rows_with_sentiment if s == 'Negative']

        for row_data in positive_rows:
            new_ws.append(row_data)

        if positive_rows and negative_rows:
            # Insert blank row separator
            new_ws.append([None] * num_cols)

        for row_data in negative_rows:
            new_ws.append(row_data)

        style_sheet(new_ws)

    new_wb.save(full_path)
    
    # Count sheets
    sheet_info = []
    for sn in new_wb.sheetnames:
        ws = new_wb[sn]
        data_rows = sum(1 for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True) if any(v is not None for v in row))
        sheet_info.append(sn + ': ' + str(data_rows))
    print(prefix + ' -> ' + ', '.join(sheet_info))

print('\nDone! All files updated with blank row separators between Positive and Negative test cases.')
